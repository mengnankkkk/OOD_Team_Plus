from pathlib import Path
import json

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "public" / "全部A股股票代码一览表.xlsx"
TARGET = ROOT / "src" / "data" / "a-share-stocks.json"


def main() -> None:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code, name = row[:2]
        if code is None or name is None:
            continue
        normalized_code = str(code).strip().split(".")[0].zfill(6)
        normalized_name = str(name).strip()
        if normalized_code and normalized_name:
            rows.append({"code": normalized_code, "name": normalized_name})
    TARGET.parent.mkdir(parents=True, exist_ok=True)
    TARGET.write_text(json.dumps(rows, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"wrote {len(rows)} rows to {TARGET}")


if __name__ == "__main__":
    main()
